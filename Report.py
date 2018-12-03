import os
import pickle
import re
import numpy as np
import pandas as pd
import urllib.request
from bs4 import BeautifulSoup
from collections import OrderedDict
from IPython.display import display
from matplotlib import pyplot as plt
from openpyxl import load_workbook


ipeds = pd.read_excel(r'Files\Databases\ipeds database.xlsx')

def get_report():
    working_files = r'Files\Working Files'
    reports = pd.Series(os.listdir(working_files))
    display(pd.DataFrame(reports))
    report = input('\nPlease choose a file: ', )

    try:
        results = reports[int(report)]
        return os.path.join(working_files, results)
    except (ValueError, KeyError):
        print('\n{0} is not a valid inuput...\nPlease choose another file: '.format(report))
        return get_report()


class CESS(object):
    def __init__(self, client_file, name=None, file_format='csv', transform=True, ):
        if file_format.lower() == 'csv':
            self.cess = pd.read_csv(client_file)
        elif file_format.lower() == 'excel':
            self.cess = pd.read_excel(client_file)
        else:
            self.cess = client_file

        self.name = name
        self.year = pd.to_datetime('today').year
        self.sample_size = self.cess.shape[0]
        self.ipeds_id = None
        self.client_type = None
        self.client_flag = True

        self.labels = {j: self.cess.iloc[0, i] for i, j in enumerate(self.cess)}

        self.scale = {'IMP': {'very important': 5,
                              'important': 4,
                              'somewhat important': 3,
                              'not very important': 2,
                              'not important at all': 1},
                      'SAT': {'very satisfied': 5,
                              'very satisfied': 5,
                              'satisfied': 4,
                              'somewhat satisfied': 3,
                              'not very satisfied': 2,
                              'not satisfied at all': 1},
                      'GOAL': {'[A] Increase the enrollment of new students': 1,
                             '[B] Retain more of its current students to graduation': 2,
                             '[C] Improve the academic ability of entering student classes': 3,
                             '[D] Recruit students from new geographic markets': 4,
                             '[E] Increase the diversity of racial and ethnic groups represented among the student body': 5,
                             '[F] Develop new academic programs': 6,
                             '[G] Improve the quality of existing academic programs': 7,
                             '[H] Improve the appearance of campus buildings and grounds': 8,
                             '[I] Improve employee morale': 9},
                      'INV': {'too much involvement': 5,
                              'more than enough involvement': 4,
                              'just the right involvement': 3,
                              'not quite enough involvement': 2,
                              'not enough involvement': 1},
                      'Q7': {'Less than 1 year': 1,
                               '1 to 5 years': 2,
                               '6 to 10 years': 3,
                               '11 to 20 years': 4,
                               'More than 20 years': 5},
                      'Q8': {'Full-time': 1,
                             'Part-time': 2},
                      'Q9': {'Faculty': 1,
                             'Staff': 2,
                             'Administrator': 3}}

        # the level ordering of self.conversions is as follows: Section, Question, Conversion
        # i.e self.conversion['Q1_IMP']['Q1_1_IMP'] = {'very important': 5...}
        self.conversions = OrderedDict((('Demographics', {}),))
        for i in ['Q7', 'Q8', 'Q9']:
            self.conversions['Demographics'][i] = OrderedDict((k, self.scale[i][k]) for k in self.scale[i])

        for question in self.cess.loc[2:, 'Q7':]:
            try:
                demo_question_dict = self.conversions['Demographics'][question]
                max_demo_count = max([demo_question_dict[k] for k in demo_question_dict])
            except KeyError:
                demo_question_dict = {j.strip(): i for i, j in enumerate(self.cess.loc[2:, :][question].value_counts().sort_index().index, 1)}
                self.conversions['Demographics'][question] = demo_question_dict
            if re.match('Q\d+', question):
                for i in self.cess.loc[2:, question].value_counts().index:
                    if i.strip() not in demo_question_dict.keys():
                        self.conversions['Demographics'][question][i] = max_demo_count
                        max_demo_count += 1

        demo_arrays = [[], []]
        demo_frames = []
        for i, j in enumerate(self.conversions['Demographics']):
            for k in self.conversions['Demographics'][j]:
                demo_arrays[0].append(j)
                demo_arrays[1].append(k)
        demo_index = pd.MultiIndex.from_tuples(list(zip(*demo_arrays)))
        self.base_frame_demo = pd.DataFrame(index=demo_index)

        for question in self.cess:

            match = re.match('(Q\d+)_\d+_([A-Z]+)', question)

            if match:
                q_num, q_cat = match.groups()
                try:
                    self.conversions[q_num + '_' + q_cat][question] = self.scale[q_cat]
                except KeyError:
                    self.conversions[q_num + '_' + q_cat] = OrderedDict(((question, self.scale[q_cat]),))

            elif re.match('Q3_\d', question):
                self.base_frame_q3 = self.cess.loc[0, :][self.conversions['Q2_IMP']].str.replace('.+- ', '').to_frame().set_index(0)
                for i, j in enumerate(self.base_frame_q3.index, 1):
                    q3_convert = OrderedDict((j, i) for i, j in enumerate(self.base_frame_q3.index, 1))
                    try:
                        self.conversions['Q3'][question] = q3_convert
                    except KeyError:
                        self.conversions['Q3'] = OrderedDict(((question, q3_convert),))

            elif re.match('Q4_\d+', question):
                try:
                    self.conversions['Q4_INV'][question] = self.scale['INV']
                except KeyError:
                    self.conversions['Q4_INV'] = OrderedDict(((question, self.scale['INV']),))

            elif re.match('Q6$', question):
                self.conversions['Q6_SAT'] = OrderedDict(((question, self.scale['SAT']),))

        self.conversions.move_to_end('Demographics')


        self.coding_key = {}
        for section in self.conversions:
            for question in self.conversions[section]:
                values = self.conversions[section][question]
                for value in values:
                    self.coding_key[value.lower()] = values[value]


    def set_name(self, name):
        self.name = name


    def get_type(self, ipeds_id):
        try:
            url = urllib.request.urlopen('http://nces.ed.gov/collegenavigator/?q=all&id={0}'.format(ipeds_id))
            soup = BeautifulSoup(url, "lxml")
            list = soup.find_all('table')[1].find_all('tr')[2].text
            wordlist = list.split()[1:len(list)]
            Client_Type = ' '.join(wordlist)
            print(ipeds_id, '\n', Client_Type)
            return Client_Type
        except IndexError:
            print("No type found for IPED's ID {0}".format(ipeds_id))
            return np.nan


    def get_id(self, name):
        potentials = ipeds[ipeds['college'].str.contains(name, case=False)].reset_index(drop=True)
        final = potentials['ipeds id'].values
        while len(final) != 1:
            display(potentials)
            new_search = input('Please choose college, enter new search IPEDS ID manually: ')
            if new_search.isdigit() and int(new_search) in potentials.index:
                return potentials.loc[int(new_search), 'ipeds id']
            elif new_search.isdigit():
                return int(new_search)
            potentials = ipeds[ipeds.college.str.match(new_search, case=False)].reset_index(drop=True)
            final = potentials['ipeds id'].values
        display(potentials['college'])
        return int(final[0])


    def descriptives(self, section, category, df=None, remove=''):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]
        return df[section].applymap(lambda x: self.scale[category][x.lower()] if x == x else x).describe().rename(columns={x: re.sub(remove, '', self.labels[x]) for x in self.labels}).T[['mean', 'std', 'count']]


    def counts(self, df=None):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]
        return self.base_frame_q3.join(df.loc[:, 'Q3_1':'Q3_3'].apply(lambda x: x.str.strip().value_counts()).sort_index(), how='left').fillna(0)


    def demographics(self, df=None):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]
        return self.base_frame_demo.join(df.loc[:, list(self.conversions['Demographics'])].apply(lambda x: x.str.strip().value_counts()).unstack().to_frame(), how='left').fillna(0)


    def open_ends(self, df=None):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]

        for i in df:
            if re.match('OE_\d+', i):
                open_ended_text = re.sub('\[QID40-QuestionText]', self.name, self.labels[i])
                final = df[[i]].dropna().rename(columns={i: open_ended_text})
                yield final


    def run_report(self, df=None, remove='.+- '):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]

        for j in self.conversions:
            if re.match('.+_[A-Z]+', j):
                match = re.match('.+_([A-Z]+)', j).groups()[0]
                yield j, self.descriptives(list(self.conversions[j]), match, df=df, remove=remove)
            elif j == 'Q3':
                yield j, self.counts(df=df)
            elif j == 'Demographics':
                yield j, self.demographics(df=df)
            else:
                pass


    def main_report(self, df=None, show=True, report='Main', custom_template=''):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :]

        check = [30, 30, 10, 3, 8, 21, 21]
        extended = [len(self.conversions[x]) for x in self.conversions][:7] != check

        if custom_template:
            custom_template = ' - ' + custom_template

        if extended:
            book = load_workbook(r'Files\Main Report Extended Template{0}.xlsx'.format(custom_template))
            print('Main Report ran with extended template used')
            for i in self.conversions:
                print(i, len(self.conversions[i]))
        else:
            book = load_workbook(r'Files\Main Report Normal Template{0}.xlsx'.format(custom_template))
            print('Main Report ran with Normal template used')

        writer = pd.ExcelWriter(r'Deliverables\{0} CESS_{1} {2} Report.xlsx'.format(self.name, self.year, report),
                                engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        row = 0
        sheet = 'Conversion Sheet'
        cell_positions = [(0, 0), (0, 6), (42, 0), (57, 0), (72, 0), (85, 0), (85, 6), (118, 0), (121, 0), (133, 0), (0, 0)]

        for frame, k in zip(self.run_report(df=df), cell_positions):
            if show:
                display(*frame)
                print('\n')
            if type(frame[1]) == pd.DataFrame:
                frame[1].to_excel(writer, sheet_name=sheet, startrow=k[0], startcol=k[1])
            for i in self.open_ends():
                row = 0
                col = 0
                for index, f in enumerate(self.open_ends(df=df), 1):
#                     f.loc[df.shape[0] + 1] = '' # not sure if needed
                    f.iloc[0:, :].to_excel(writer, sheet_name='Open-ended item {0}'.format(index), index=False,
                                           startrow=row, startcol=col)
        writer.save()


    def raw_data(self):
        raw_data_columns = []
        for i in self.conversions:
            for j in list(self.conversions[i]):
                raw_data_columns.append(j)

        text_data = self.cess.loc[2:, :][raw_data_columns]
        coded_data = self.cess.loc[2:, :][raw_data_columns].applymap(lambda x: self.coding_key[x.lower().strip()] if x == x else x)

        coding_key = pd.DataFrame()
        for section in self.conversions:
            for col in self.conversions[section]:
                reversed_key = {self.conversions[section][col][x]:x for x in self.conversions[section][col]}
                question = pd.concat([pd.DataFrame({col: reversed_key}).rename(columns={col: 'Label'})], keys=[col])
                coding_key = pd.concat([coding_key, question], axis=0)

        # Create variable label row
        label_header = ['StartDate', 'EndDate', 'Finished']
        for label in text_data:
            label_header.extend([re.sub('.+- ', '', self.labels[label])])
        var_labels = pd.DataFrame(index=label_header).T

        # Create text and coded raw data sheets
        text = self.cess[['StartDate', 'EndDate', 'Finished']].iloc[2:, :].join(text_data)
        coded = self.cess[['StartDate', 'EndDate', 'Finished']].iloc[2:, :].join(coded_data)

        # Create an new Excel file and add worksheets.
        writer = pd.ExcelWriter(r'Deliverables\Raw Data.xlsx', engine='xlsxwriter')

        # add text worksheet.
        var_labels.to_excel(writer, sheet_name='Raw Data', startrow=1, index=False)
        text.to_excel(writer, sheet_name='Raw Data', startrow=2, index=False)

        # add coded worksheet.
        var_labels.to_excel(writer, sheet_name='Raw Data Coded', startrow=1, index=False)
        coded.to_excel(writer, sheet_name='Raw Data Coded', startrow=2, index=False)

        # add coding key worksheet.
        coding_key.to_excel(writer, sheet_name='Coding Key', startrow=1)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        raw_worksheet = writer.sheets['Raw Data']
        coded_worksheet = writer.sheets['Raw Data Coded']

        header = ["Survey Introduction", "SECTION 1: Campus culture and policies",
                  "SECTION 1: Campus culture and policies continued"
            , "SECTION 2: Institutional goals", "SECTION 2: Institutional goals (continued)",
                  "3: Involvement in planning and decision-making"
            , "SECTION 4: Work environment", "SECTION 4: Work environment continued",
                  "SECTION 4: Work environment (continued)"
            , "SECTION 5: Demographics"]

        start_count = 3
        raw_worksheet.merge_range(0, 0, 0, 2, header[0])
        coded_worksheet.merge_range(0, 0, 0, 2, header[0])
        for section, head in zip(self.conversions, header[1:]):
            if section != 'Q6_SAT':
                raw_worksheet.merge_range(0, start_count, 0, start_count + len(self.conversions[section]) - 1, head)
                coded_worksheet.merge_range(0, start_count, 0, start_count + len(self.conversions[section]) - 1, head)
            else:
                raw_worksheet.write_string(0, start_count + len(self.conversions[section]) - 1, head)
                coded_worksheet.write_string(0, start_count + len(self.conversions[section]) - 1, head)
            start_count += len(self.conversions[section])

        # format coding key sheet
        cd_worksheet = writer.sheets['Coding Key']

        # sets up format to use in the specified range.
        cell_format = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#d9d9d9'})

        # adds the cell formatting to specified cells
        cd_worksheet.merge_range('A1:C1', 'Variable Values', cell_format)
        cd_worksheet.merge_range('A2:B2', 'Values', cell_format)
        cd_worksheet.write('C2', 'Values', cell_format)
        writer.save()

        return text, coded, coding_key


    def filter_frame(self, f_question=[], f_operand=[], f_value=[], f_bool=[]):
        df = self.cess.loc[2:, :]
        f_bool.append('')
        transformed_values = []
        for question, value in zip(f_question, f_value):
            for transformed_value in self.conversions['Demographics'][question]:
                if value == self.conversions['Demographics'][question][transformed_value]:
                    transformed_values.append(transformed_value)
        report_type = '-'.join(transformed_values)

        filter_string = ''.join([str(z[0]) + str(z[1]) + "'{0}'".format(z[2]) + str(z[3]) for z in zip(f_question, f_operand, transformed_values, f_bool)])
        transformed_string = filter_string
        filtered_frame = self.cess.loc[df.applymap(lambda x: str(x).strip()).query(filter_string).index]
        for i in f_question:
            transformed_string = re.sub(i, self.cess.loc[0, i], transformed_string)
        print('{0}\nSegment logic: {1}\n\nRecords left: {2}'.format(report_type, transformed_string, filtered_frame.shape[0]))
        return filtered_frame, report_type


    def segment_report(self, report=None, f_question=[], f_operand=[], f_value=[], f_bool=[]):
        df, report_type = self.filter_frame(f_question=f_question,
                                f_operand=f_operand, f_value=f_value, f_bool=f_bool)

        if df.empty:
            return 'Segment contains no records'

        if not report:
            report = report_type

        return self.main_report(df=df, report=report)


    def set_comparison_variables(self, name=None):
        if name is None:
            name = self.name
        self.ipeds_id = self.get_id(name)
        self.client_type = self.get_type(self.ipeds_id)
        self.cess['Client_name'] = self.name
        self.cess['Year'] = pd.to_datetime('today').year
        self.cess['Sample_size'] = self.cess.shape[0]
        self.cess['Client_IPED_ID'] = self.ipeds_id
        self.cess['Client_Type'] = self.client_type
        self.cess['Client_Flag'] = True


    def keep_type(self, c_type, filter_ipeds=None):
        if c_type == '2-year, Public':
            print('Comparison to 2-year, Public schools')
            filtered_df = self.cess[self.cess.Client_Type.isin(['2-year, Public'])]
        elif c_type == '4-year, Private not-for-profit':
            print('Comparison to 4-year, Private not-for-profit schools')
            filtered_df = self.cess[self.cess.Client_Type.isin(['4-year, Private not-for-profit'])]
        elif c_type in ("2-year, Private for-profit", '2-year, Private not-for-profit'):
            print('Comparison to all 2-year schools')
            filtered_df = self.cess[self.cess.Client_Type.isin(
                ("2-year, Private for-profit", '2-year, Private not-for-profit', '2-year, Public'))]
        elif c_type in ('4-year, Private', '4-year, Public', "4-year, primarily associate's, Private for-profit"
                      , "4-year, primarily associate's, Private not-for-profit",
                      "4-year, primarily associate's, Public", "4-year, Private not-for-profit"):
            print('Comparison to all 4 year schools')
            filtered_df = self.cess[self.cess.Client_Type.isin(('4-year, Private not-for-profit', '4-year, Private',
                                                              '4-year, Public',
                                                              "4-year, primarily associate's, Private for-profit"
                                                              , "4-year, primarily associate's, Private not-for-profit",
                                                              "4-year, primarily associate's, Public"))]
        else:
            print('Shool not found. Used 4-year, Private not-for-profit')
            filtered_df = self.cess[self.cess.Client_Type.isin(['4-year, Private not-for-profit'])]
        if (filtered_df.Client_IPED_ID == filter_ipeds).sum() > 0:
            print(' & '.join([x for x in filtered_df[filtered_df.Client_IPED_ID == filter_ipeds]['Client_name'].value_counts().index]) + ' have been filtered out')
            filtered_df = filtered_df[filtered_df.Client_IPED_ID != filter_ipeds]

        return filtered_df


    def statistical_report(self, df=None, other_cess=None, other_frame=None, show=True, report='Comparison', group1=None, group2=None):
        if type(df) != pd.DataFrame:
            df = self.cess.loc[2:, :'Q9']
    #     if type(other_frame) != pd.DataFrame:
    #         with open('Master File.pickle', 'rb') as f:
    #             master = pickle.load(f)

        if report == 'Statistical':
            file_name = '{0} Comparison to {1}'.format(group1, group2)
            cell_positions = [(2, 0), (46, 0), (90, 0), (107, 0), (127, 0), (142, 0), (176, 0), (211, 0), (216, 0), (233, 0)]
        else:
            file_name = 'Comparison'
            cell_positions = [(2, 0), (36, 0), (70, 0), (83, 0), (103, 0), (115, 0), (140, 0), (165, 0), (170, 0)]

        book = load_workbook(r'Files\{0} Report TEMPLATE.xlsx'.format(report))
        writer = pd.ExcelWriter(r'Deliverables\{0} CESS_2018 {1} Report.xlsx'.format(self.name, file_name),
                                engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        row = 0
        sheet = 'Conversion Sheet'
        for main_frame, comparison_frame, k in zip(self.run_report(df=df), other_cess.run_report(df=other_frame), cell_positions):
            frame = main_frame[1].iloc[:, :3].join(comparison_frame[1], how='inner', rsuffix='_c')
            if show:
                display(frame)
                print('\n')
            if type(frame) == pd.DataFrame:
                frame.to_excel(writer, sheet_name=sheet, startrow=k[0], startcol=k[1])

        if report == 'Statistical':
            pd.DataFrame(index=[group1, group2]).T.to_excel(writer, sheet_name=sheet, startrow=230, startcol=0)
        else:
            pd.DataFrame([[self.cess.Client_name.value_counts().index[0]]]).to_excel(writer, sheet_name=sheet, startrow=188,
                                                                                     startcol=0)
            pd.DataFrame(other_frame.Client_name.value_counts().sort_index()).to_excel(writer, sheet_name=sheet,
                                                                                       startrow=192, startcol=0)
        writer.save()


    def statistically_tested_report(self, segment1=None, segment2=None
                                    , g1_question=[], g1_operand=[], g1_value=[], g1_bool=[]
                                    , g2_question=[], g2_operand=[], g2_value=[], g2_bool=[]):
        df, group1 = self.filter_frame(f_question=g1_question, f_operand=g1_operand, f_value=g1_value, f_bool=g1_bool)
        df2, group2 = self.filter_frame(f_question=g2_question, f_operand=g2_operand, f_value=g2_value, f_bool=g2_bool)

        if segment1 and segment2:
            group1, group2 = segment1, segment2

        self.statistical_report(df=df, other_cess=self, other_frame=df2, report='Statistical', group1=group1, group2=group2)


    def comparison_report(self):
        with open(r'files\Databases\Master Files\Master CESS.pickle', 'rb') as f:
            master = pickle.load(f)

        self.set_comparison_variables()
        filtered_master = master.keep_type(self.client_type, filter_ipeds=str(self.ipeds_id))
        self.statistical_report(df=self.cess.loc[2:, :], other_cess=master, other_frame=filtered_master)

        keep_columns = [x for x in master.cess.columns if x in self.cess.columns]
        master.cess = pd.concat([master.cess, self.cess.loc[2:, :][keep_columns]], sort=False).reset_index(drop=True)

        save_master = input('Do you want to save the master file (y,n)...')
        if save_master == 'y':
            with open(r'files\Databases\Master Files\Master CESS.pickle', 'wb') as f:
                pickle.dump(master, f, protocol=pickle.HIGHEST_PROTOCOL)
